import csv
from pathlib import Path
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import tkinter as tk

agency = Path("Agency")
agency.mkdir(exist_ok=True)

templates = agency / "Templates"
templates.mkdir(exist_ok=True)

contract_template = templates / "Contract_Template.pdf"
with open(contract_template, "w") as pdf_file:
    pdf_file.write("This is the Template to be used for Future Contracts")


class Client_App:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Agency Client Management System")
        self.window.geometry("600x650")

        title = tk.Label(
            self.window,
            text="AGENCY CLIENT MANAGEMENT SYSTEM\nv1",
            font=("Arial", 18, "bold")
        )
        title.pack(pady=15)


        name_label = tk.Label(self.window, text="Name")
        name_label.pack()
        self.name_entry = tk.Entry(self.window, width=40)
        self.name_entry.pack(pady=5)

        email_label = tk.Label(self.window, text="Email")
        email_label.pack()
        self.email_entry = tk.Entry(self.window, width=40)
        self.email_entry.pack(pady=5)

        company_label = tk.Label(self.window, text="Company")
        company_label.pack()
        self.company_entry = tk.Entry(self.window, width=40)
        self.company_entry.pack(pady=5)

        project_label = tk.Label(self.window, text="Project")
        project_label.pack()
        self.project_entry = tk.Entry(self.window, width=40)
        self.project_entry.pack(pady=5)

        amount_label = tk.Label(self.window, text="Amount")
        amount_label.pack()
        self.amount_entry = tk.Entry(self.window, width=40)
        self.amount_entry.pack(pady=5)

        status_label = tk.Label(self.window, text="Status")
        status_label.pack()
        self.status_entry = tk.Entry(self.window, width=40)
        self.status_entry.pack(pady=5)


        onboard_button = tk.Button(self.window, text="Onboard Client", command=self.add_client, width=25)
        onboard_button.pack(pady=10)

        view_all_button = tk.Button(self.window, text="View All Leads", command=self.view_all_leads, width=25)
        view_all_button.pack()

        view_new_button = tk.Button(self.window, text="View New Leads", command=self.view_new_leads, width=25)
        view_new_button.pack(pady=5)

        exit_button = tk.Button(self.window, text="Exit", command=self.window.destroy, width=25)
        exit_button.pack(pady=10)

        results_label = tk.Label(self.window, text="Results")
        results_label.pack()
        self.result = tk.Listbox(self.window, width=60, height=10)
        self.result.pack(pady=10)

        self.search_button = tk.Button(self.window, text="Search Lead", command=self.search_lead, width=25)
        self.search_button.pack(pady=10)
        self.search_entry = tk.Entry(self.window, width=40)
        self.search_entry.pack()

        self.window.mainloop()


    def add_client(self):
        self.result.delete(0, tk.END)

        name = self.name_entry.get()
        email = self.email_entry.get()
        company = self.company_entry.get()
        project = self.project_entry.get()
        amount = self.amount_entry.get()
        status = self.status_entry.get()

        if name == "" or company == "":
            self.result.insert(tk.END,"Please fill in all the required fields.")
            return

        client = agency / company
        client.mkdir(exist_ok=True)

        contracts  = client / "Contracts"
        contracts.mkdir(exist_ok=True)

        assets = client / "Assets"
        assets.mkdir(exist_ok=True)

        deliverables = client / "Deliverables"
        deliverables.mkdir(exist_ok=True)

        reports = client / "Reports"
        reports.mkdir(exist_ok=True)

        shutil.copy(contract_template, contracts)

        workbook = load_workbook("Client_Tracker.xlsx")
        sheet = workbook.active

        new_row = sheet.max_row + 1

        sheet.cell(row=new_row, column=1).value = name
        sheet.cell(row=new_row, column=2).value = email
        sheet.cell(row=new_row, column=3).value = company
        sheet.cell(row=new_row, column=4).value = project
        sheet.cell(row=new_row, column=5).value = amount
        sheet.cell(row=new_row, column=6).value = status

        for cell in ["A1", "B1", "C1", "D1", "E1", "F1"]:
            sheet[cell].font = Font(bold=True)
            sheet[cell].alignment = Alignment(horizontal="center", vertical="center")

        workbook.save("Client_Tracker.xlsx")

        with open("leads.csv", "a", newline="") as csvfile:
            leads = csv.writer(csvfile)
            leads.writerow([name, email, company, project, amount, status])

        self.result.insert(tk.END, "✓ Client Folder Created")
        self.result.insert(tk.END, "✓ Contract Copied")
        self.result.insert(tk.END, "✓ Excel Updated")
        self.result.insert(tk.END, "✓ Leads Database Updated")
        self.result.insert(tk.END, "")
        self.result.insert(tk.END, "Client Successfully Onboarded!")

        self.name_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.company_entry.delete(0, tk.END)
        self.project_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)
        self.status_entry.delete(0, tk.END)

        self.name_entry.focus()


    def view_all_leads(self):
        self.result.delete(0, tk.END)

        self.result.insert(tk.END, "====================")
        self.result.insert(tk.END, "      ALL LEADS     ")
        self.result.insert(tk.END, "====================")

        with open("leads.csv", "r") as csvfile:
            leads = csv.Dictreader(csvfile)
            for row in leads:
                self.result.insert(tk.END, f"{row['name']} | {row['company']} | {row['status']}")


    def view_new_leads(self):
        self.result.delete(0, tk.END)
        self.result.insert(tk.END, "===================")
        self.result.insert(tk.END, "     NEW LEADS     ")
        self.result.insert(tk.END, "===================")

        with open("leads.csv", "r") as csvfile:
            leads = csv.DictReader(csvfile)
            for row in leads:
                if row['status'] == "New":
                    self.result.insert(tk.END, f"{row['name']} | {row['company']}")

    def search_lead(self):
        self.result.delete(0, tk.END)

        company = self.search_entry.get()

        with open("leads.csv", "r") as csvfile:
            leads = csv.DictReader(csvfile)
            found = False

            for row in leads:
                if row['company'].lower() == company.lower():
                 self.result.insert(tk.END, f"{row['name']} | {row['company']}")

                 found = True

            if not found:
                self.result.insert(tk.END, "No company found.")

app = Client_App
app()














